"""공인 위생검사전문기관(식품·축산물 시험·검사기관) DB 및 검증.

자가품질검사성적서를 발급한 검사기관이 식약처가 지정한 공인 시험·검사기관 목록에
실제로 존재하는지 확인한다. 목록은 chandra/data/test_agencies.json 에 보관하며,
식약처가 배포하는 '위생검사전문기관 목록' PDF 의 텍스트 레이어를
``build_agency_db_from_pdf`` 로 결정적 파싱해 채운다(식품만, 축산물 제외).
"""

from __future__ import annotations

import json
import re
from datetime import date as _date
from dataclasses import asdict, dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from chandra.text_match import digits as _digits, strip_entity as _norm

_DB_PATH = Path(__file__).with_name("data") / "test_agencies.json"


@dataclass
class TestAgency:
    name: str
    designation_no: str | None = None  # 지정번호 (예: 식품 제099호)
    category: str | None = None  # 식품 / 축산물
    address: str | None = None
    tel: str | None = None  # 대표 전화번호 (OCR 강건 매칭 키)
    representative: str | None = None  # 대표자명
    work_scope: str | None = None  # 업무범위 (자가품질위탁검사/전문/수입 등)
    field_detail: str | None = None  # 분야 상세 (식품/건강기능식품/첨가물 등)
    scopes: list[str] = field(default_factory=list)  # 시험·검사항목 (이화학/미생물 등)
    aliases: list[str] = field(default_factory=list)
    valid_until: str | None = None  # 지정 유효기간

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class AgencyVerification:
    queried_name: str | None
    queried_designation_no: str | None
    found: bool
    matched: TestAgency | None
    match_basis: str | None  # designation_no / name / alias
    designation_no_match: bool | None
    detail: str
    designation_expired: bool | None = None  # 지정 유효기간 경과 여부(True=만료)

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        data["matched"] = self.matched.to_dict() if self.matched else None
        return data


def _parse_agency_date(text: str | None) -> _date | None:
    """'26.3.9' / '2026-03-09' / '2026.3.9' 등 검사기관 유효기간 표기를 date 로.

    2자리 연도는 2000년대로 본다.
    """
    if not text:
        return None
    m = re.search(r"(\d{2,4})[.\-/년]\s*(\d{1,2})[.\-/월]\s*(\d{1,2})", str(text))
    if not m:
        return None
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100:
        y += 2000
    try:
        return _date(y, mo, d)
    except ValueError:
        return None


def _desig_category(text: str | None) -> str | None:
    """지정번호 표기에서 분야 추출 — '식품 제099호'→식품, '축산물 제26호'→축산물."""
    if not text:
        return None
    if "축산" in text:
        return "축산물"
    if "식품" in text:
        return "식품"
    return None


def _norm_designation(text: str | None) -> str:
    """'식품 제145호'/'제145호'/'제099호' → 숫자만('145','99')으로 정규화.

    DB는 분야 접두 없이 '제N호'로, 성적서는 '식품 제N호'로 표기될 수 있어 숫자만 비교한다.
    분야 구분은 dedup 키의 category 와 verify_agency 의 이름/전화 매칭이 담당한다.
    """
    if not text:
        return ""
    m = re.search(r"제?\s*0*(\d+)\s*호", text)
    if m:
        return str(int(m.group(1)))
    return "".join(text.split())


def load_agency_db(path: str | Path = _DB_PATH) -> list[TestAgency]:
    p = Path(path)
    if not p.exists():
        return []
    data = json.loads(p.read_text(encoding="utf-8"))
    return [
        TestAgency(
            name=row.get("name", ""),
            designation_no=row.get("designation_no"),
            category=row.get("category"),
            address=row.get("address"),
            tel=row.get("tel"),
            representative=row.get("representative"),
            work_scope=row.get("work_scope"),
            field_detail=row.get("field_detail"),
            scopes=row.get("scopes", []) or [],
            aliases=row.get("aliases", []) or [],
            valid_until=row.get("valid_until"),
        )
        for row in data.get("agencies", [])
    ]


def verify_agency(
    name: str | None,
    designation_no: str | None = None,
    tel: str | None = None,
    address: str | None = None,
    db: list[TestAgency] | None = None,
    today: _date | None = None,
) -> AgencyVerification:
    """검사기관이 공인 목록에 있는지 확인.

    매칭 우선순위: 전화번호(숫자) > 지정번호 > 기관명/별칭 > 퍼지(유사도).
    전화번호는 OCR에서 숫자라 잘 안 깨지므로 가장 강건한 키다.
    지정번호 매칭은 분야(식품/축산물)가 어긋나면 인정하지 않는다(식품 제99호 ≠ 축산물 제26호).
    매칭된 기관의 지정 유효기간이 today 를 지났으면 designation_expired=True 로 표시한다.
    """
    agencies = db if db is not None else load_agency_db()
    ref = today or _date.today()
    nname = _norm(name)
    ndesig = _norm_designation(designation_no)
    ntel = _digits(tel)
    want_cat = _desig_category(designation_no)

    def _mk(found, ag, basis, desig_ok, detail):
        expired = None
        if ag and ag.valid_until:
            d = _parse_agency_date(ag.valid_until)
            if d:
                expired = d < ref
                if expired:
                    detail += f" — ⚠️ 지정 유효기간 만료(~{d.isoformat()}), 적합 인정 불가"
        return AgencyVerification(
            name, designation_no, found, ag, basis, desig_ok, detail, designation_expired=expired
        )

    # 0) 전화번호(숫자) 매칭 — 성적서 하단 전화번호로 식별 (가장 강건)
    #    오탐 방지를 위해 전체 자릿수 동일 또는 9자리 이상 접미 일치만 인정.
    if len(ntel) >= 9:
        for ag in agencies:
            atel = _digits(ag.tel)
            if not atel or len(atel) < 9:
                continue
            if ntel == atel or ntel[-10:] == atel[-10:] or ntel[-9:] == atel[-9:]:
                if want_cat and ag.category and ag.category != want_cat:
                    continue  # 분야(식품/축산물)가 다르면 전화 접미 우연일치로 보고 건너뜀
                return _mk(
                    True, ag, "tel",
                    None if not ndesig else _norm_designation(ag.designation_no) == ndesig,
                    f"전화번호({tel})로 공인기관 매칭: {ag.name}",
                )

    # 1) 지정번호 매칭 (분야 일치 요구 — 식품/축산물 교차매칭 방지)
    if ndesig:
        for ag in agencies:
            if _norm_designation(ag.designation_no) != ndesig:
                continue
            if want_cat and ag.category and ag.category != want_cat:
                continue  # 같은 번호라도 분야가 다르면 다른 기관
            return _mk(
                True, ag, "designation_no", True,
                f"지정번호 '{designation_no}' 로 공인기관 매칭: {ag.name}",
            )

    # 2) 기관명/별칭 정확·부분 매칭
    if nname:
        for ag in agencies:
            keys = [ag.name, *ag.aliases]
            if want_cat and ag.category and ag.category != want_cat:
                continue  # 분야가 다른 기관은 이름이 비슷해도 제외
            for key in keys:
                nk = _norm(key)
                if nk and (nk == nname or nk in nname or nname in nk):
                    desig_ok = (
                        None
                        if not ndesig
                        else _norm_designation(ag.designation_no) == ndesig
                    )
                    basis = "name" if nk == _norm(ag.name) else "alias"
                    detail = f"기관명으로 공인기관 매칭: {ag.name}"
                    if desig_ok is False:
                        detail += (
                            f" (단, 지정번호 불일치: 성적서 '{designation_no}' vs "
                            f"DB '{ag.designation_no}')"
                        )
                    return _mk(True, ag, basis, desig_ok, detail)

    # 3) 퍼지 매칭 (OCR 글자 오인식 보정: 디아이→다이아이 등) — 기관명 + 주소 보조
    if nname:
        naddr = _norm(address)
        best_ag, best_score, best_basis = None, 0.0, "fuzzy"
        for ag in agencies:
            name_score = max(
                (SequenceMatcher(None, nname, _norm(k)).ratio() for k in [ag.name, *ag.aliases]),
                default=0.0,
            )
            # 주소 유사도가 높으면 가산 (전화/지정번호가 없을 때 보조 식별)
            addr_score = (
                SequenceMatcher(None, naddr, _norm(ag.address)).ratio()
                if naddr and ag.address
                else 0.0
            )
            score = max(name_score, 0.6 * name_score + 0.4 * addr_score)
            if score > best_score:
                best_ag, best_score = ag, score
                best_basis = "fuzzy+addr" if addr_score > name_score else "fuzzy"
        if best_ag and best_score >= 0.7:
            return _mk(
                True, best_ag, best_basis,
                None if not ndesig else _norm_designation(best_ag.designation_no) == ndesig,
                f"유사도 {best_score:.2f}로 공인기관 추정 매칭: {best_ag.name} "
                f"(성적서 표기 '{name}', OCR 오인식 가능 — 확인 권장)",
            )

    return AgencyVerification(
        name, designation_no, False, None, None, None,
        "공인 위생검사전문기관 목록에서 찾지 못함 (목록 갱신 필요 여부 확인)",
    )


# ---------------------------------------------------------------------------
# 목록 PDF → DB 인제스트 (텍스트 레이어 결정적 파싱)
# ---------------------------------------------------------------------------

_REGION = (
    r"(서울|부산|대구|인천|광주|대전|울산|세종|경기|강원|충북|충남|전북|전남|"
    r"경북|경남|제주|충청|전라|경상)"
)
_SCOPE_KW = [
    "이화학", "미생물", "방사능", "잔류농약", "잔류동물용의약품",
    "유전자변형", "노로바이러스", "다이옥신", "식품조사처리",
]
_ENTRY_START = re.compile(r"^(\d+)\s+(제\s*\d+호)(.*)$")


def _is_skip_line(s: str) -> bool:
    if re.fullmatch(r"\d+", s):
        return True
    for kw in [
        "식품등의 자가품질검사 매뉴얼", "연번 지정번호", "개소",
        "위생검사전문기관", "업무범위", "시험·검사기관",
    ]:
        if kw in s:
            return True
    return False


def parse_agency_text(
    text: str,
    default_category: str = "식품",
    section_break_kw: str | None = "축산물",
) -> list[dict[str, Any]]:
    """위생검사전문기관 목록 표의 텍스트 레이어를 결정적으로 파싱한다.

    행 시작은 '연번 제NNN호'. ☎ 전화번호와 끝의 유효기간 날짜를 앵커로 사용하고,
    소재지는 행정구역 접두로 분리한다. section_break_kw 가 줄에 나오면 중단한다
    (식품 목록 파싱 시 '축산물' 섹션 시작에서 멈추기 위함; 축산물 파싱 시 None).
    """
    blocks: list[dict[str, str]] = []
    cur: dict[str, str] | None = None
    for raw in text.splitlines():
        s = raw.strip()
        if not s or _is_skip_line(s):
            continue
        if section_break_kw and section_break_kw in s:
            break
        m = _ENTRY_START.match(s)
        if m:
            if cur:
                blocks.append(cur)
            cur = {"desig": re.sub(r"\s", "", m.group(2)), "rest": m.group(3).strip()}
        elif cur is not None:
            cur["rest"] += " " + s
    if cur:
        blocks.append(cur)

    rows: list[dict[str, Any]] = []
    for b in blocks:
        blob = b["rest"]
        tel = None
        mt = re.search(r"☎\s*([0-9][0-9\-~,\s]*)", blob)
        if mt:
            tel = re.split(r"[~,]", mt.group(1).strip())[0].strip().rstrip("-")
        dates = re.findall(r"\d{4}\s*\.\s*\d{1,2}\s*\.\s*\d{1,2}\.?", blob)
        valid = re.sub(r"\s", "", dates[-1]) if dates else None
        pre = blob.split("☎")[0].strip()
        # 소재지 시작 = '행정구역(+시/군/구)' 패턴. 기관명에 포함된 지역어(예: 전라북도
        # 생물산업진흥원)는 시/군/구가 곧바로 따라오지 않으므로 건너뛴다.
        addr_re = re.compile(_REGION + r"[가-힣]*\s*\S{0,10}?(시|군|구)\b")
        mr = addr_re.search(pre)
        if not mr:
            mr = re.search(_REGION, pre)
        if mr:
            name = re.sub(r"\s+", " ", pre[: mr.start()]).strip(" ,")
            address = re.sub(r"\s+", " ", pre[mr.start():]).strip()
        else:
            name = re.sub(r"\s+", " ", pre).strip()
            address = ""
        tail = blob.split("fax")[-1]
        scopes = [k for k in _SCOPE_KW if k in tail] or [k for k in _SCOPE_KW if k in blob]
        rows.append(
            {
                "designation_no": b["desig"],
                "name": name,
                "address": address,
                "tel": tel,
                "category": default_category,
                "scopes": scopes,
                "valid_until": valid,
            }
        )
    return rows


def _sheet_category(sheet_name: str) -> str:
    if "축산물" in sheet_name:
        return "축산물"
    if "위생용품" in sheet_name:
        return "위생용품"
    return "식품"  # 식품전문 / 식품 등 자가품질위탁


def _split_scopes(text: str | None) -> list[str]:
    if not text:
        return []
    return [s.strip() for s in re.split(r"[,/、]", str(text)) if s.strip()]


def build_agency_db_from_excel(
    path: str,
    out_path: str | Path = _DB_PATH,
    csv_path: str | Path | None = None,
) -> dict[str, Any]:
    """식약처 '시험검사기관 현황' 엑셀(다중 시트)을 읽어 DB(JSON, +CSV)로 만든다.

    시트별 컬럼: 연번·지정번호·기관명·대표자명·소재지·전화·팩스·업무범위·분야·시험검사항목·유효기간.
    같은 분야 내 동일 지정번호는 병합(업무범위/항목 보강). PDF 파싱본을 대체한다(최신·정확).
    """
    import csv as _csv

    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    by_key: dict[tuple[str, str], dict[str, Any]] = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        category = _sheet_category(sheet)
        rows = list(ws.iter_rows(values_only=True))
        # 헤더 행 = '지정번호'를 포함한 행
        hidx = next(
            (i for i, r in enumerate(rows) if r and any("지정번호" == str(c).strip() for c in r if c)),
            None,
        )
        if hidx is None:
            continue
        header = [str(c).strip() if c is not None else "" for c in rows[hidx]]
        col = {name: header.index(name) for name in header if name}

        def g(r, name):
            i = col.get(name)
            v = r[i] if (i is not None and i < len(r)) else None
            return str(v).strip() if v is not None else None

        for r in rows[hidx + 1 :]:
            if not r or not any(r):
                continue
            desig = g(r, "지정번호")
            name = g(r, "기관명")
            if not desig and not name:
                continue
            rec = {
                "designation_no": desig,
                "name": name,
                "representative": g(r, "대표자명"),
                "address": g(r, "소재지"),
                "tel": g(r, "전화"),
                "work_scope": g(r, "업무범위"),
                "field_detail": g(r, "분야"),
                "scopes": _split_scopes(g(r, "시험검사항목")),
                "valid_until": g(r, "유효기간"),
                "category": category,
            }
            k = _dedup_key(rec)
            if k[1]:
                by_key.setdefault(k, {}).update({kk: v for kk, v in rec.items() if v})
    wb.close()

    agencies = list(by_key.values())
    cats = sorted({a.get("category", "") for a in agencies})
    payload = {
        "_meta": {
            "description": f"식약처 시험검사기관 현황 (분야: {', '.join(cats)})",
            "source": f"parsed from Excel: {Path(path).name}",
            "count": len(agencies),
        },
        "agencies": agencies,
    }
    Path(out_path).write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    if csv_path:
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            w = _csv.writer(f)
            w.writerow(["분야", "지정번호", "기관명", "대표자", "소재지", "전화", "업무범위", "시험검사항목", "유효기간"])
            for a in agencies:
                w.writerow([
                    a.get("category", ""), a.get("designation_no", ""), a.get("name", ""),
                    a.get("representative", ""), a.get("address", ""), a.get("tel", ""),
                    a.get("work_scope", ""), ", ".join(a.get("scopes", [])), a.get("valid_until", ""),
                ])
    return {"count": len(agencies), "by_category": {c: sum(1 for a in agencies if a.get("category") == c) for c in cats}, "json": str(out_path)}


def _dedup_key(row: dict[str, Any]) -> tuple[str, str]:
    """분야+지정번호(없으면 기관명)로 구분. 식품 제045호 ≠ 축산물 제45호 충돌 방지."""
    ident = _norm_designation(row.get("designation_no")) or _norm(row.get("name"))
    return (row.get("category") or "", ident)


def build_agency_db_from_pdf(
    path: str,
    pages: list[int],
    out_path: str | Path = _DB_PATH,
    csv_path: str | Path | None = None,
    category: str = "식품",
    section_break_kw: str | None = "축산물",
    merge: bool = False,
) -> dict[str, Any]:
    """PDF 텍스트 레이어에서 위생검사기관 목록을 파싱해 DB(JSON, +CSV)로 저장/병합.

    category: 분야('식품'/'축산물'). section_break_kw: 다른 섹션 시작에서 중단(식품='축산물').
    merge=True 면 기존 DB에 (분야,지정번호) 기준으로 합친다.
    """
    import csv as _csv

    texts = extract_pdf_text(path, pages, max_pages=len(pages) + max(pages) + 1)
    rows = parse_agency_text(
        "\n".join(texts), default_category=category, section_break_kw=section_break_kw
    )
    by_key: dict[tuple[str, str], dict[str, Any]] = {}
    out = Path(out_path)
    if merge and out.exists():
        prev = json.loads(out.read_text(encoding="utf-8"))
        for r in prev.get("agencies", []):
            by_key[_dedup_key(r)] = dict(r)
    for r in rows:
        k = _dedup_key(r)
        if k[1]:
            by_key.setdefault(k, {}).update({kk: v for kk, v in r.items() if v})
    agencies = list(by_key.values())
    cats = sorted({a.get("category", "") for a in agencies})
    payload = {
        "_meta": {
            "description": f"위생검사전문기관 목록 (분야: {', '.join(cats)})",
            "source": f"parsed from PDF text layer: {Path(path).name}",
            "count": len(agencies),
        },
        "agencies": agencies,
    }
    Path(out_path).write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    if csv_path:
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            w = _csv.writer(f)
            w.writerow(["지정번호", "기관명", "소재지", "전화번호", "분야", "시험·검사항목", "유효기간"])
            for a in agencies:
                w.writerow([
                    a.get("designation_no", ""), a.get("name", ""), a.get("address", ""),
                    a.get("tel", ""), a.get("category", ""),
                    ", ".join(a.get("scopes", [])), a.get("valid_until", ""),
                ])
    return {"count": len(agencies), "json": str(out_path), "csv": str(csv_path) if csv_path else None}


def extract_pdf_text(path: str, pages: list[int] | None, max_pages: int) -> list[str]:
    """PDF 페이지의 텍스트 레이어를 추출한다(페이지별 문자열)."""
    import pypdfium2 as pdfium

    doc = pdfium.PdfDocument(path)
    try:
        if pages is None:
            indices = range(min(len(doc), max_pages))
        else:
            indices = [i for i in pages if 0 <= i < len(doc)]
        return [doc[i].get_textpage().get_text_range() for i in indices]
    finally:
        doc.close()
